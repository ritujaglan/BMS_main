import pickle
import smtplib
from openpyxl import Workbook,load_workbook
import datetime
import os
import pathlib
import sys


class Account:
    accNo = 0
    name = ''
    email=''
    deposit = 0
    type = ''
    count=2

    @classmethod
    def acct_counter(cls):
        cls.count += 1
        return cls.count

    def create_account(self):
        try:
            self.accNo = int(input("Enter the account no : "))
            self.name = input("Enter the account holder name : ")
            self.type = input("Enter the type of account [C/S] : ")
            self.email = input("Enter the account holder email: ")
            self.deposit = int(input("Enter The Initial amount(>=500 for Saving and >=1000 for current: "))
            print("\n\n\nAccount Created")
        except Exception as e:
            print(str(e))
            print('Please rerun the program and try with correct value')
            sys.exit(1)

    def write_excel_file(self):
        count=Account.acct_counter()
        print('This is at counter:',Account.count)
        file1 = pathlib.Path("AcctDetails.xlsx")
        if file1.exists():
            workbook = load_workbook(filename='AcctDetails.xlsx')
            sheet = workbook.active
            row_count=sheet.max_row
            B = 'B' + str(row_count+1)
            C = 'C' + str(row_count+1)
            D = 'D' + str(row_count+1)
            E = 'E' + str(row_count+1)
            F = 'F' + str(row_count+1)
            sheet[B] = str(self.accNo)
            sheet[C] = self.name
            sheet[D] = self.deposit
            sheet[E] = self.type
            sheet[F] = datetime.date.today()
            workbook.save(filename='AcctDetails.xlsx')
        else:
            workbook=Workbook()
            sheet=workbook.active
            sheet['A1']='Sno.'
            sheet['B1'] = 'AcctNo'
            sheet['C1'] = 'Acctname'
            sheet['D1'] = 'Acct_Balance'
            sheet['E1'] = 'Acct_type'
            sheet['F1'] = 'Acct_Opening_Date'
            sheet['B2'] = str(self.accNo)
            sheet['C2'] = self.name
            sheet['D2'] = self.deposit
            sheet['E2'] = self.type
            sheet['F2'] = datetime.date.today()
            workbook.save(filename='AcctDetails.xlsx')
        print('File opening......')
        os.startfile('AcctDetails.xlsx')


    def show_account(self):
        print("Account Number : ", self.accNo)
        print("Account Holder Name : ", self.name)
        print("Type of Account", self.type)
        print("Balance : ", self.deposit)
        print("email : ", self.email)

    # def modifyAccount(self):
    #     print("Account Number : ", self.accNo)
    #     self.name = input("Modify Account Holder Name :")
    #     self.type = input("Modify type of Account :")
    #     self.deposit = int(input("Modify Balance :"))

    def deposit_amount(self, amount):
        self.deposit += amount

    def withdraw_amount(self, amount):
        self.deposit -= amount

    def report(self):
        print(self.accNo, " ", self.name, " ", self.type, " ", self.deposit)

    def get_accountno(self):
        return self.accNo

    def get_account_holder_name(self):
        return self.name

    def get_account_type(self):
        return self.type

    def get_deposit(self):
        return self.deposit


def intro():
    print("\t\t\t\t**********************")
    print("\t\t\t\tMY BANK")
    print("Enter anything to start")
    input()


def write_account():
    account = Account()
    account.create_account()
    account.write_excel_file()
    write_accounts_file(account)


def display_all():
    file = pathlib.Path("accounts.data")
    if file.exists():
        infile = open('accounts.data', 'rb')
        mylist = pickle.load(infile)
        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = 'Sno.'
        sheet['B1'] = 'AcctNo.'
        sheet['C1'] = 'Customer Name'
        sheet['D1'] = 'Acct Balance'
        sheet['E1'] = 'Acct Type'
        sheet['F1'] = 'Date'
        workbook.save(filename='AcctDetails.xlsx')
        c=2
        for item in mylist:
            print("ACCT No.:",item.accNo, '\n'"ACCT Holder's Name.: ", item.name, ", ACCT Type: ", item.type, ", ACCT Balance: ", item.deposit)
            #A='A'+str(c+1)
            B='B'+str(c)
            C='C'+str(c)
            D='D'+str(c)
            E = 'E' + str(c)
            F = 'F' + str(c)
            sheet[B] = str(item.accNo)
            sheet[C] = item.name
            sheet[D] = item.deposit
            sheet[E] = item.type
            sheet[F] = datetime.datetime.today()
            workbook.save(filename='AcctDetails.xlsx')
            c+=1
        infile.close()
        print('File Opening.......')
        os.startfile('AcctDetails.xlsx')
        print('File closing.......')
    else:
        print("No records to display")


def display_sp(num):
    file = pathlib.Path("accounts.data")
    if file.exists():
        infile = open('accounts.data', 'rb')
        mylist = pickle.load(infile)
        infile.close()
        found = False
        for item in mylist:
            if item.accNo == num:
                print("Your account Balance is = ", item.deposit)
                found = True
    else:
        print("No records to Search")
    if not found:
        print("No existing record with this number")

def email_alert(num1,num2,email,amount,deposit):
    from1 = "test12feb21@gmail.com"
    password='Test@1234'
    # from1 = "ritu.jaglan2008@gmail.com"
    # password = 'dilliabhidurhai1'
    # to=email
    subject='Your Account Update Satus'
    email_text=""
    if num2==1:
        #email_text='Rs.'+ amount +' Amount Deposited in your account and your current balance is: '
        email_text = f'Rs. {amount} Amount Deposited in your account and your current balance is {deposit} '
    else:
        email_text = f'Rs. {amount} Amount withdrawn from your account and your current balance is {deposit} '
    try:
        server=smtplib.SMTP_SSL('smtp.gmail.com',587)
        server.ehlo()
        #server.starttls()
        server.login(from1,password)
        server.sendmail(from1,to,email_text)
        server.close()
        print('Email sent')
    except Exception as e:
        #print(e)
        print('Something went wrong while trying to send mail to user with the below error:')
        print(e)

def deposit_and_withdraw(num1, num2):
    file = pathlib.Path("accounts.data")
    if file.exists():
        infile = open('accounts.data', 'rb')
        mylist = pickle.load(infile)
        infile.close()
        os.remove('accounts.data')
        found=False
        for item in mylist:
            if item.accNo == num1:
                found=True
                if num2 == 1:
                    amount = int(input("Enter the amount to deposit : "))
                    item.deposit += amount
                    print("Your account is updated and your current balance is: ",item.deposit)
                    email_alert(num1,num2,item.email,amount,item.deposit)
                elif num2 == 2:
                    amount = int(input("Enter the amount to withdraw : "))
                    if amount <= item.deposit:
                        item.deposit -= amount
                        print("Your account is updated and your current balance is: ", item.deposit)
                        email_alert(num1,num2, item.email,amount,item.deposit)
                    else:
                        print("You cannot withdraw larger amount")
        if found is False:
            print(f'\t Account No. {num1} does not exists')
    else:
        print("No records to Search")
    outfile = open('newaccounts.data', 'wb')
    pickle.dump(mylist, outfile)
    outfile.close()
    os.rename('newaccounts.data', 'accounts.data')

def write_accounts_file(account):
    file = pathlib.Path("accounts.data")
    if file.exists():
        infile = open('accounts.data', 'rb')
        oldlist = pickle.load(infile)
        oldlist.append(account)
        infile.close()
        os.remove('accounts.data')
    else:
        oldlist = [account]
    outfile = open('newaccounts.data', 'wb')
    pickle.dump(oldlist, outfile)
    outfile.close()
    os.rename('newaccounts.data', 'accounts.data')

def delete_content():
    file1=open('accounts.data','r+')
    file1.seek(0)
    file1.truncate(0)
    file1.close()
    #file1.seek(0) # I believe this seek is redundant
    #return file1

#tempFile=deleteContent(tempFile)

# start of the program
ch = ''
num = 0
intro()

while ch != 8:
    # system("cls");
    print("\nMAIN MENU")
    print("\t1. NEW ACCOUNT")
    print("\t2. DEPOSIT AMOUNT")
    print("\t3. WITHDRAW AMOUNT")
    print("\t4. BALANCE ENQUIRY")
    print("\t5. ALL ACCOUNT HOLDER LIST")
    # print("\t6. CLOSE AN ACCOUNT")
    # print("\t7. MODIFY AN ACCOUNT")
    print("\t8. EXIT")
    print("\tSelect Your Option (1-8) ")
    ch = input()
    # system("cls");

    if ch == '1':
        write_account()
    elif ch == '2':
        num = int(input("\tEnter The account No. : "))
        deposit_and_withdraw(num, 1)
    elif ch == '3':
        num = int(input("\tEnter The account No. : "))
        deposit_and_withdraw(num, 2)
    elif ch == '4':
        num = int(input("\tEnter The account No. : "))
        display_sp(num)
    elif ch == '5':
        display_all()
    elif ch == '8':
        print("\tBye")
        break
    else:
        print("Invalid choice")
    ch = input("\nPlease enter anything you want to run more options else press n to exit")
    if ch=='n':
        sys.exit(0)














